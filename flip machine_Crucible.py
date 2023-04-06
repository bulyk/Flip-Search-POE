import cloudscraper
import time
import requests
from openpyxl import load_workbook
import tkinter as tk
from tkinter.ttk import Label

timeout = 12

league = 'Standard'

HEADERS = {'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.160 YaBrowser/22.5.2.615 Yowser/2.5 Safari/537.36'}

headers = {
    'accept': 'application/json',
    'user-agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/100.0.4896.160 YaBrowser/22.5.2.615 Yowser/2.5 Safari/537.36',

    # Already added when you pass json=
    # 'Content-Type': 'application/json',
}
nomer_items = int(3)

def pricer_div(divine_price):
    # card = 'The Sephirot'
    json_data_card = {
        'query': {
            'filters': {
                'trade_filters': {
                    'disabled': False,
                #     'filters': {
                #         'price': {
                #             'min': 1,
                #             'max': 30,
                #         },
                #     },
                },
            },
            'status': {
                'option': 'online',
            },
            'stats': [
                {
                    'type': 'and',
                    'filters': [],
                },
            ],
            'Type': card,
        },
        'sort': {
            'price': 'asc',
        },
    }
    response_div = requests.post('https://www.pathofexile.com/api/trade/search/'+ league, headers=headers, json=json_data_card, timeout=2)
    print(response_div)
    try: 
        result_id = response_div.json()['result'][nomer_items]

        scraper = cloudscraper.create_scraper(delay=25)
        fetch = 'https://www.pathofexile.com/api/trade/fetch/'+result_id
        requests_fetch = scraper.get(fetch)
        price_card_chaos_or_ex = requests_fetch.json()['result'][0]['listing']['price']['currency']
    #    print(price_card_chaos_or_ex)
        if 'chaos' in price_card_chaos_or_ex:
            price_card = requests_fetch.json()['result'][0]['listing']['price']['amount']
        elif 'divine' in price_card_chaos_or_ex:
            price_card = requests_fetch.json()['result'][0]['listing']['price']['amount']*divine_price
    #    print(price_card)
        time.sleep(timeout)
        price_item(divine_price, price_card)
    except:
        time.sleep(timeout)
        price_card = 'нет карточек!'
        print('нет карточек!')

type_item = {
    0: 'name',
    1:  'Type'      
             }
n = 0
def price_item(divine_price, price_card):
    global n
    if 'Divine' not in item:
        json_data_item = {
            'query': {
                'filters': {
                    'trade_filters': {
                        'disabled': False,
                    #     'filters': {
                    #         'price': {
                    #             'min': 1,
                    #             'max': 30,
                    #         },
                    #     },
                    },
                },
                'status': {
                    'option': 'online',
                },
                'stats': [
                    {
                        'type': 'and',
                        'filters': [],
                    },
                ],
                type_item[n]: item,
            },
            'sort': {
                'price': 'asc',
            },
        }
        response_item = requests.post('https://www.pathofexile.com/api/trade/search/'+ league, headers=headers, json=json_data_item, timeout=2)

        try: 
            result_id_item = response_item.json()['result'][nomer_items]

            scraper = cloudscraper.create_scraper(delay=25)
            fetch_item = 'https://www.pathofexile.com/api/trade/fetch/'+result_id_item
            requests_fetch_item = scraper.get(fetch_item)
            price_item_chaos_or_ex = requests_fetch_item.json()['result'][0]['listing']['price']['currency']
        #    print(price_item_chaos_or_ex)
            if 'chaos' in price_item_chaos_or_ex:
                price_item = requests_fetch_item.json()['result'][0]['listing']['price']['amount']
            elif 'divine' in price_item_chaos_or_ex:
                price_item = requests_fetch_item.json()['result'][0]['listing']['price']['amount']*divine_price
        #    print(price_item)
            time.sleep(timeout)
        except:
            time.sleep(timeout)
            price_item = 'нет предмета'
            print('нет предмета')



    else:
        price_item = divine_price
######################################
    profit = int(price_item)*int(quantity_item) - int(price_card)*int(quantity_div)
    print (card, profit)
    fn = 'table.xlsx'
    wb = load_workbook(fn)
    ws = wb['trade']
    ws['G2'] = divine_price
    ws.append([card, price_card, int(quantity_div), profit, item, quantity_item])
    wb.save(fn)
    wb.close()

##################################
def divine():
    json_data_divine = {
        'query': {
            'filters': {
                'trade_filters': {
                    'disabled': False,
                #     'filters': {
                #         'price': {
                #             'min': 1,
                #             'max': 30,
                #         },
                #     },
                },
            },
            'status': {
                'option': 'online',
            },
            'stats': [
                {
                    'type': 'and',
                    'filters': [],
                },
            ],
            'type': 'Divine Orb',
        },
        'sort': {
            'price': 'asc',
        },
    }
    response_divine = requests.post('https://www.pathofexile.com/api/trade/search/'+ league, headers=headers, json=json_data_divine, timeout=2)
    print(response_divine)
    global nomer_items
    try: 
        
        result_id_divine = response_divine.json()['result'][nomer_items]

        scraper = cloudscraper.create_scraper(delay=25)
        fetch_divine = 'https://www.pathofexile.com/api/trade/fetch/'+result_id_divine
        requests_fetch_divine = scraper.get(fetch_divine)
        price_divine_chaos_or_ex = requests_fetch_divine.json()['result'][0]['listing']['price']['currency']
        #    print(price_card_chaos_or_ex)
        if 'chaos' in price_divine_chaos_or_ex:
            divine_price = requests_fetch_divine.json()['result'][0]['listing']['price']['amount']
            time.sleep(timeout)
            pricer_div(divine_price)
        else:
            nomer_items += 1
            time.sleep(timeout)
            divine()
    except:
        time.sleep(timeout)
        print('диваны не продают?')


if __name__ == "__main__":
    
    with open(r"div_flip_list.txt", 'r', encoding='utf-8') as fp:
        len_zap = len(fp.readlines())
        print('Total Number of lines:', len_zap)

    root = tk.Tk()
    root.geometry('300x200')
    root.resizable(False, False)
    root.title('Flip machine')
    x = 0
    file = open('div_flip_list.txt', 'r', encoding='utf-8')
    lines = file.readlines()

    label = tk.Label(root, text="")
    label.pack()

    for line in lines:
        last_line = line.strip()
        obrez = last_line.split('-')
        quantity_div = obrez[0].split(' (')[1].split(')')[0].lstrip().rstrip()
        card = obrez[0].split('(')[0].lstrip().rstrip()
        item = obrez[1].split('(')[0].lstrip().rstrip()
        quantity_item = obrez[1].split('(')[1].split(')')[0].lstrip().rstrip()


        label.config(text=str(x)+'/'+str(len_zap))
        root.update()

        # label = Label(root, text=str(x)+'/'+str(len_zap))
        label.pack(ipadx=10, ipady=10)
        root.update()
        divine()
        x += 1
    root.mainloop(len_zap)


